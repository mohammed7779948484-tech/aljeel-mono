# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import io
import json
import math
from datetime import datetime
from pathlib import Path
from typing import Any

import frappe
import requests
from openpyxl import load_workbook
from frappe.utils import get_datetime, now
from frappe.utils.file_manager import save_file
from xlrd import open_workbook

from .registry import ADMIN_ROLES
from .utils import ApiError, api_endpoint, require_roles


TOP_K = 4
MIN_SIMILARITY = 0.35
SEMANTIC_WEIGHT = 0.8
KEYWORD_WEIGHT = 0.2
MAX_EMBED_RECORDS = 1000
EMPTY_QUESTION_MESSAGE = "اكتب سؤالك من فضلك."
NO_INFO_MESSAGE = "حاليًا لا أجد إجابة مباشرة لهذا السؤال ضمن بيانات الموقع المتاحة لدي."
OUTSIDE_MESSAGE = "أعتذر، أنا مساعد مخصص لأسئلة جامعة الجيل الجديد فقط. اسألني عن القبول، الكليات، البرامج، الرسوم، أو خدمات الجامعة."
GENERIC_ERROR_MESSAGE = "حدث خلل مؤقت. حاول مرة أخرى بعد قليل."

QUESTION_HEADERS = {
    "question",
    "question_ar",
    "question_arabic",
    "q",
    "faq_question",
    "السؤال",
    "سؤال",
    "questionen",
    "question_en",
}
ANSWER_HEADERS = {
    "answer",
    "answer_ar",
    "answer_arabic",
    "a",
    "faq_answer",
    "الاجابة",
    "الإجابة",
    "الجواب",
    "جواب",
    "answeren",
    "answer_en",
}


def _normalize_arabic(text: str) -> str:
    return (
        str(text or "")
        .replace("أ", "ا")
        .replace("إ", "ا")
        .replace("آ", "ا")
        .replace("ة", "ه")
        .replace("ى", "ي")
        .replace("ـ", "")
        .strip()
        .lower()
    )


def _normalize_question(text: str) -> str:
    q = str(text or "").strip()
    slang_map = {
        "فين": "اين",
        "وين": "اين",
        "ايش": "ما",
        "اش": "ما",
        "شلون": "كيف",
        "قديش": "كم",
    }
    for src, target in slang_map.items():
        q = q.replace(src, target)
    return " ".join(_normalize_arabic(q).split())


def _classify_question(question: str) -> str:
    q = _normalize_question(question)
    small_talk = {
        "السلام عليكم",
        "السلام",
        "مرحبا",
        "اهلا",
        "ما اسمك",
        "من انت",
        "مين انت",
        "كيف استخدم الشات",
        "ساعدني",
    }
    if any(token in q for token in small_talk):
        return "small_talk"

    university_terms = {
        "جامعه",
        "الجامعه",
        "الجيل الجديد",
        "قبول",
        "تسجيل",
        "رسوم",
        "منح",
        "كليه",
        "الكليات",
        "برنامج",
        "برامج",
        "تخصص",
        "الموقع",
        "العنوان",
        "تواصل",
        "البريد",
        "الهاتف",
        "شؤون الطلاب",
        "الدوام",
        "المواعيد",
        "اين",
    }
    if any(token in q for token in university_terms):
        return "university"

    return "outside"


def _small_talk_reply(question: str, doc) -> str:
    q = _normalize_question(question)
    if any(token in q for token in ["السلام عليكم", "السلام", "مرحبا", "اهلا"]):
        return f"وعليكم السلام! أنا {doc.chat_name_ar or 'المساعد الذكي'} وأساعدك بمعلومات جامعة الجيل الجديد."
    if any(token in q for token in ["ما اسمك", "من انت", "مين انت"]):
        return f"أنا {doc.chat_name_ar or 'المساعد الذكي'} الخاص بجامعة الجيل الجديد."
    if any(token in q for token in ["كيف استخدم الشات", "ساعدني"]):
        return "اكتب سؤالك عن القبول، الكليات، البرامج، الرسوم، أو خدمات الجامعة وسأجيبك من قاعدة المعرفة المتاحة."
    return f"مرحبًا بك! أنا {doc.chat_name_ar or 'المساعد الذكي'}."


def _cosine_similarity(a: list[float], b: list[float]) -> float:
    dot = 0.0
    norm_a = 0.0
    norm_b = 0.0
    for x, y in zip(a, b):
        dot += x * y
        norm_a += x * x
        norm_b += y * y
    if not norm_a or not norm_b:
        return 0.0
    return dot / (math.sqrt(norm_a) * math.sqrt(norm_b))


def _keyword_score(question: str, faq_question: str) -> float:
    q_words = [w for w in question.split(" ") if w]
    if not q_words:
        return 0.0
    faq_words = set(w for w in faq_question.split(" ") if w)
    overlap = sum(1 for word in q_words if word in faq_words)
    return overlap / len(q_words)


def _top_matches(question_embedding: list[float] | None, normalized_question: str, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    matches = []
    for item in items:
        item_embedding = item.get("embedding") or []
        semantic_score = _cosine_similarity(question_embedding or [], item_embedding) if question_embedding and item_embedding else 0.0
        keyword_score = _keyword_score(normalized_question, _normalize_question(item.get("question") or ""))
        final_score = semantic_score * SEMANTIC_WEIGHT + keyword_score * KEYWORD_WEIGHT
        matches.append(
            {
                **item,
                "semanticScore": semantic_score,
                "keywordScore": keyword_score,
                "finalScore": final_score,
            }
        )
    return sorted(matches, key=lambda row: row["finalScore"], reverse=True)[:TOP_K]


def _parse_response(response: requests.Response) -> dict[str, Any]:
    try:
        return response.json()
    except Exception:
        return {}


def _provider_defaults(provider: str) -> tuple[str, str]:
    if provider == "OpenAI":
        return "gpt-4.1-mini", "text-embedding-3-small"
    return "gemini-2.5-flash", "gemini-embedding-001"


def _get_file_doc_by_url(file_url: str):
    if not file_url:
        return None
    file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
    if not file_name:
        return None
    return frappe.get_doc("File", file_name)


def _resolve_site_file_path(file_url: str) -> Path:
    normalized = str(file_url or "").split("?")[0].strip()
    if not normalized:
        raise ApiError("VALIDATION_ERROR", "Knowledge source file is missing.", status_code=400)
    if normalized.startswith("/private/files/"):
        relative = normalized[len("/private/files/") :]
        return Path(frappe.get_site_path("private", "files", relative))
    if normalized.startswith("/files/"):
        relative = normalized[len("/files/") :]
        return Path(frappe.get_site_path("public", "files", relative))
    return Path(frappe.get_site_path(normalized.lstrip("/")))


def _read_file_bytes(file_url: str) -> bytes:
    file_doc = _get_file_doc_by_url(file_url)
    if file_doc:
        content = file_doc.get_content()
        return content if isinstance(content, bytes) else str(content or "").encode("utf-8")

    physical_path = _resolve_site_file_path(file_url)
    if not physical_path.exists():
        raise ApiError(
            "VALIDATION_ERROR",
            "Knowledge source file is not saved on the server yet. Save the form again and verify the file upload.",
            status_code=400,
        )
    return physical_path.read_bytes()


def _get_settings_doc():
    if not frappe.db.exists("DocType", "Smart Chat Settings"):
        raise ApiError("NOT_CONFIGURED", "Smart Chat Settings is not installed.", status_code=500)
    return frappe.get_single("Smart Chat Settings")


def _get_api_key(doc) -> str:
    api_key = doc.get_password("provider_api_key", raise_exception=False)
    if not api_key:
        raise ApiError("VALIDATION_ERROR", "Provider API key is missing.", status_code=400)
    return api_key


def _embed_text(provider: str, api_key: str, model: str, text: str) -> list[float]:
    if provider == "OpenAI":
        response = requests.post(
            "https://api.openai.com/v1/embeddings",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={"model": model, "input": text},
            timeout=60,
        )
        payload = _parse_response(response)
        if not response.ok:
            raise ApiError("PROVIDER_ERROR", f"OpenAI embedding failed: {payload or response.text}", status_code=502)
        values = payload.get("data", [{}])[0].get("embedding")
        if not isinstance(values, list) or not values:
            raise ApiError("PROVIDER_ERROR", "OpenAI embedding response is invalid.", status_code=502)
        return values

    response = requests.post(
        f"https://generativelanguage.googleapis.com/v1beta/models/{model}:embedContent",
        params={"key": api_key},
        headers={"Content-Type": "application/json"},
        json={"content": {"parts": [{"text": text}]}},
        timeout=60,
    )
    payload = _parse_response(response)
    if not response.ok:
        raise ApiError("PROVIDER_ERROR", f"Gemini embedding failed: {payload or response.text}", status_code=502)
    values = ((payload or {}).get("embedding") or {}).get("values")
    if not isinstance(values, list) or not values:
        raise ApiError("PROVIDER_ERROR", "Gemini embedding response is invalid.", status_code=502)
    return values


def _generate_answer(provider: str, api_key: str, model: str, temperature: float, question: str, matches: list[dict[str, Any]]) -> str:
    context = "\n\n".join(
        [f"Q{i + 1}: {item.get('question')}\nA{i + 1}: {item.get('answer')}" for i, item in enumerate(matches)]
    )
    prompt = "\n".join(
        [
            "أنت مساعد رسمي لجامعة الجيل الجديد.",
            "أجب بالعربية فقط وباختصار وبدقة.",
            "أجب فقط من المصدر.",
            f"إذا لم تكن الإجابة موجودة بوضوح في المصدر فقل فقط: {NO_INFO_MESSAGE}",
            "",
            f"السؤال:\n{question}",
            "",
            f"المصدر:\n{context}",
        ]
    )

    if provider == "OpenAI":
        response = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "model": model,
                "temperature": temperature,
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=90,
        )
        payload = _parse_response(response)
        if not response.ok:
            raise ApiError("PROVIDER_ERROR", f"OpenAI generation failed: {payload or response.text}", status_code=502)
        text = (((payload or {}).get("choices") or [{}])[0].get("message") or {}).get("content")
        return str(text or NO_INFO_MESSAGE).strip()

    response = requests.post(
        f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent",
        params={"key": api_key},
        headers={"Content-Type": "application/json"},
        json={
            "contents": [{"role": "user", "parts": [{"text": prompt}]}],
            "generationConfig": {"temperature": temperature},
        },
        timeout=90,
    )
    payload = _parse_response(response)
    if not response.ok:
        raise ApiError("PROVIDER_ERROR", f"Gemini generation failed: {payload or response.text}", status_code=502)
    text = (((((payload or {}).get("candidates") or [{}])[0].get("content") or {}).get("parts") or [{}])[0].get("text"))
    return str(text or NO_INFO_MESSAGE).strip()


def _load_rows_from_source(file_url: str) -> list[list[Any]]:
    if not file_url:
        raise ApiError("VALIDATION_ERROR", "Knowledge source file is missing.", status_code=400)

    ext = Path(file_url.split("?")[0]).suffix.lower()
    if ext == ".csv":
        content = _read_file_bytes(file_url).decode("utf-8-sig", errors="ignore")
        sample = content[:4096]
        delimiter = ","
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            delimiter = dialect.delimiter
        except Exception:
            first_line = sample.splitlines()[0] if sample.splitlines() else ""
            if first_line.count(";") > first_line.count(","):
                delimiter = ";"
        return list(csv.reader(io.StringIO(content, newline=""), delimiter=delimiter, quotechar='"'))
    if ext == ".xlsx":
        workbook = load_workbook(io.BytesIO(_read_file_bytes(file_url)), read_only=True, data_only=True)
        sheet = workbook.active
        return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    if ext == ".xls":
        workbook = open_workbook(file_contents=_read_file_bytes(file_url))
        sheet = workbook.sheet_by_index(0)
        return [sheet.row_values(idx) for idx in range(sheet.nrows)]

    raise ApiError("VALIDATION_ERROR", "Only xlsx, xls, and csv files are supported.", status_code=400)


def _sanitize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _rows_to_entries(rows: list[list[Any]]) -> list[dict[str, str]]:
    if not rows or len(rows) < 2:
        raise ApiError("VALIDATION_ERROR", "The uploaded file does not contain enough rows.", status_code=400)

    header = [_sanitize_header(cell) for cell in rows[0]]
    question_indexes = [idx for idx, name in enumerate(header) if name in QUESTION_HEADERS]
    answer_indexes = [idx for idx, name in enumerate(header) if name in ANSWER_HEADERS]

    if not question_indexes or not answer_indexes:
        raise ApiError(
            "VALIDATION_ERROR",
            "The file must contain question and answer columns such as question/answer or سؤال/الإجابة.",
            status_code=400,
        )

    entries: list[dict[str, str]] = []
    for row in rows[1:]:
        question = ""
        answer = ""
        for idx in question_indexes:
            if idx < len(row) and str(row[idx] or "").strip():
                question = str(row[idx]).strip()
                break
        for idx in answer_indexes:
            if idx < len(row) and str(row[idx] or "").strip():
                answer = str(row[idx]).strip()
                break
        if question and answer:
            entries.append({"question": question, "answer": answer})

    if not entries:
        raise ApiError("VALIDATION_ERROR", "No valid question-answer rows were found in the file.", status_code=400)

    return entries


def _delete_existing_file(file_url: str | None) -> None:
    if not file_url:
        return
    file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
    if file_name:
        frappe.delete_doc("File", file_name, ignore_permissions=True, force=True)


def _load_index(file_url: str) -> dict[str, Any]:
    if not file_url:
        raise ApiError("NOT_CONFIGURED", "Knowledge index file is missing. Rebuild the chat index first.", status_code=400)
    content = _read_file_bytes(file_url).decode("utf-8", errors="ignore")
    payload = json.loads(content)
    if not isinstance(payload, dict) or not isinstance(payload.get("items"), list) or not payload.get("items"):
        raise ApiError("VALIDATION_ERROR", "Knowledge index file is invalid.", status_code=400)
    return payload


@frappe.whitelist(allow_guest=True)
@api_endpoint
def get_public_smartchat_config():
    doc = _get_settings_doc()
    return {
        "enabled": bool(doc.enabled),
        "chatNameAr": doc.chat_name_ar or "المساعد الذكي",
        "chatNameEn": doc.chat_name_en or "Smart Assistant",
        "provider": doc.provider or "Gemini",
    }


@frappe.whitelist()
@api_endpoint
def test_smartchat_provider():
    require_roles(ADMIN_ROLES)
    doc = _get_settings_doc()
    provider = doc.provider or "Gemini"
    generation_default, embedding_default = _provider_defaults(provider)
    api_key = _get_api_key(doc)
    _embed_text(provider, api_key, doc.embedding_model or embedding_default, "اختبار ربط الشات")
    return {
        "provider": provider,
        "generationModel": doc.generation_model or generation_default,
        "embeddingModel": doc.embedding_model or embedding_default,
        "ok": True,
    }


@frappe.whitelist()
@api_endpoint
def rebuild_smartchat_index():
    require_roles(ADMIN_ROLES)
    doc = _get_settings_doc()
    provider = doc.provider or "Gemini"
    generation_default, embedding_default = _provider_defaults(provider)
    api_key = _get_api_key(doc)
    rows = _load_rows_from_source(doc.knowledge_source_file)
    entries = _rows_to_entries(rows)

    use_embeddings = len(entries) <= MAX_EMBED_RECORDS
    embedding_error = None
    items = []
    for idx, entry in enumerate(entries, start=1):
        embedding: list[float] = []
        if use_embeddings:
            try:
                embedding = _embed_text(
                    provider,
                    api_key,
                    doc.embedding_model or embedding_default,
                    _normalize_question(entry["question"]) or entry["question"],
                )
            except ApiError as exc:
                # Why: large FAQ files or temporary provider quotas should not block index creation.
                use_embeddings = False
                embedding_error = exc.message
                embedding = []
        items.append(
            {
                "id": idx,
                "question": entry["question"],
                "answer": entry["answer"],
                "embedding": embedding,
            }
        )

    payload = {
        "createdAt": now(),
        "count": len(items),
        "provider": provider,
        "generationModel": doc.generation_model or generation_default,
        "embeddingModel": doc.embedding_model or embedding_default,
        "searchMode": "semantic" if use_embeddings else "keyword",
        "embeddingSkippedReason": embedding_error
        or (f"Embedding was skipped because the file contains more than {MAX_EMBED_RECORDS} rows." if len(entries) > MAX_EMBED_RECORDS else None),
        "items": items,
    }
    content = json.dumps(payload, ensure_ascii=False)

    old_file = doc.knowledge_index_file
    saved = save_file(
        "smart_chat_index.json",
        content.encode("utf-8"),
        doc.doctype,
        doc.name,
        is_private=1,
    )
    doc.knowledge_index_file = saved.file_url
    doc.indexed_records = len(items)
    doc.last_indexed_on = get_datetime(datetime.utcnow())
    doc.save(ignore_permissions=True)
    if old_file and old_file != saved.file_url:
        _delete_existing_file(old_file)
    frappe.db.commit()

    return {
        "ok": True,
        "count": len(items),
        "indexFile": saved.file_url,
        "provider": provider,
        "searchMode": payload["searchMode"],
        "embeddingSkippedReason": payload["embeddingSkippedReason"],
    }


@frappe.whitelist()
@api_endpoint
def inspect_smartchat_source_file():
    require_roles(ADMIN_ROLES)
    doc = _get_settings_doc()
    file_url = doc.knowledge_source_file
    if not file_url:
        return {
            "ok": False,
            "hasFieldValue": False,
            "hasFileRow": False,
            "existsOnServer": False,
            "fileUrl": None,
            "serverPath": None,
            "message": "Knowledge source file is not saved in Smart Chat Settings.",
        }

    file_doc = _get_file_doc_by_url(file_url)
    physical_path = _resolve_site_file_path(file_url)
    return {
        "ok": bool(file_doc or physical_path.exists()),
        "hasFieldValue": True,
        "hasFileRow": bool(file_doc),
        "existsOnServer": physical_path.exists(),
        "fileUrl": file_url,
        "serverPath": str(physical_path),
        "message": "Knowledge source file inspection completed.",
    }


@frappe.whitelist()
@api_endpoint
def inspect_smartchat_source_rows():
    require_roles(ADMIN_ROLES)
    doc = _get_settings_doc()
    rows = _load_rows_from_source(doc.knowledge_source_file)
    entries = _rows_to_entries(rows)
    return {
        "rowCount": len(rows),
        "entryCount": len(entries),
        "header": rows[0] if rows else [],
        "sampleRows": rows[:3],
        "sampleEntries": entries[:3],
    }


@frappe.whitelist()
@api_endpoint
def inspect_smartchat_index_save():
    require_roles(ADMIN_ROLES)
    doc = _get_settings_doc()
    payload = {"createdAt": now(), "count": 1, "searchMode": "keyword", "items": [{"id": 1, "question": "test", "answer": "test", "embedding": []}]}
    content = json.dumps(payload, ensure_ascii=False)
    saved = save_file("smart_chat_index_probe.json", content.encode("utf-8"), doc.doctype, doc.name, is_private=1)
    probe_url = saved.file_url
    physical_path = _resolve_site_file_path(probe_url)
    result = {
        "probeFileUrl": probe_url,
        "probeExistsOnServer": physical_path.exists(),
        "canAssignToDoc": False,
        "canSaveDoc": False,
    }
    try:
        old_file = doc.knowledge_index_file
        doc.knowledge_index_file = probe_url
        doc.indexed_records = 1
        doc.last_indexed_on = get_datetime(datetime.utcnow())
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        result["canAssignToDoc"] = True
        result["canSaveDoc"] = True
        if old_file and old_file != probe_url:
            _delete_existing_file(old_file)
    finally:
        if probe_url and frappe.db.get_value("File", {"file_url": probe_url}, "name"):
            _delete_existing_file(probe_url)
    return result


@frappe.whitelist(allow_guest=True)
@api_endpoint
def ask_smart_chat(question: str | None = None):
    doc = _get_settings_doc()
    if not doc.enabled:
        return {"answer": GENERIC_ERROR_MESSAGE, "type": "university", "suggestions": []}

    question = (question or frappe.form_dict.get("question") or "").strip()
    if not question:
        return {"answer": EMPTY_QUESTION_MESSAGE, "type": "small_talk", "suggestions": []}

    question_type = _classify_question(question)
    if question_type == "small_talk":
        return {"answer": _small_talk_reply(question, doc), "type": question_type, "suggestions": []}
    if question_type == "outside":
        return {"answer": OUTSIDE_MESSAGE, "type": question_type, "suggestions": []}

    provider = doc.provider or "Gemini"
    generation_default, embedding_default = _provider_defaults(provider)
    api_key = _get_api_key(doc)
    index = _load_index(doc.knowledge_index_file)

    normalized_question = _normalize_question(question)
    question_embedding = None
    if index.get("searchMode") == "semantic":
        try:
            question_embedding = _embed_text(provider, api_key, doc.embedding_model or embedding_default, normalized_question or question)
        except ApiError:
            question_embedding = None
    matches = _top_matches(question_embedding, normalized_question, index["items"])
    suggestions = [item.get("question") for item in matches[:3] if item.get("question")]
    top_score = matches[0]["finalScore"] if matches else 0

    if top_score < MIN_SIMILARITY:
        return {
            "answer": "\n".join(
                [
                    f"{NO_INFO_MESSAGE} لأساعدك بشكل أدق، هل تقصد:",
                    "- القبول والتسجيل؟",
                    "- الرسوم والمنح؟",
                    "- البرامج والتخصصات؟",
                    "- التواصل والعنوان؟",
                ]
            ),
            "type": "university",
            "suggestions": suggestions,
        }

    try:
        answer = _generate_answer(
            provider,
            api_key,
            doc.generation_model or generation_default,
            float(doc.temperature or 0.25),
            question,
            matches,
        )
    except ApiError:
        # Why: if generation quota is temporarily hit, return the best matched answer directly.
        answer = matches[0].get("answer") if matches else NO_INFO_MESSAGE
    return {
        "answer": answer or NO_INFO_MESSAGE,
        "type": "university",
        "suggestions": suggestions,
    }
